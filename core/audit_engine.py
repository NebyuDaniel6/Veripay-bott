"""
Audit Engine for VeriPay - Statement reconciliation and report generation
"""
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
import yaml
from loguru import logger
import os
from pathlib import Path
import json
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class AuditEngine:
    """Audit engine for statement reconciliation and reporting"""
    
    def __init__(self, config_path: str = "config.yaml"):
        """Initialize audit engine"""
        with open(config_path, 'r') as file:
            self.config = yaml.safe_load(file)
        
        self.audit_config = self.config['audit']
        self.report_format = self.audit_config['report_format']
        
        # Create reports directory
        self.reports_dir = Path("reports")
        self.reports_dir.mkdir(exist_ok=True)
    
    def reconcile_statements(self, transactions: List[Dict], 
                           bank_statement_path: str, 
                           bank_type: str) -> Dict:
        """
        Reconcile waiter transactions with bank statement
        
        Args:
            transactions: List of transaction dictionaries
            bank_statement_path: Path to bank statement file
            bank_type: Type of bank (cbe, telebirr, dashen)
            
        Returns:
            Dict containing reconciliation results
        """
        try:
            # Parse bank statement
            statement_data = self._parse_bank_statement(bank_statement_path, bank_type)
            
            if not statement_data:
                return {
                    'success': False,
                    'error': 'Could not parse bank statement',
                    'reconciliation_results': None
                }
            
            # Convert transactions to DataFrame
            transactions_df = pd.DataFrame(transactions)
            
            # Perform reconciliation
            reconciliation_results = self._perform_reconciliation(
                transactions_df, statement_data, bank_type
            )
            
            # Generate summary statistics
            summary = self._generate_reconciliation_summary(reconciliation_results)
            
            return {
                'success': True,
                'reconciliation_results': reconciliation_results,
                'summary': summary,
                'statement_data': statement_data
            }
            
        except Exception as e:
            logger.error(f"Error in statement reconciliation: {e}")
            return {
                'success': False,
                'error': str(e),
                'reconciliation_results': None
            }
    
    def _parse_bank_statement(self, file_path: str, bank_type: str) -> Optional[pd.DataFrame]:
        """Parse bank statement file (Excel, CSV, PDF)"""
        try:
            file_extension = Path(file_path).suffix.lower()
            
            if file_extension == '.xlsx' or file_extension == '.xls':
                return self._parse_excel_statement(file_path, bank_type)
            elif file_extension == '.csv':
                return self._parse_csv_statement(file_path, bank_type)
            elif file_extension == '.pdf':
                return self._parse_pdf_statement(file_path, bank_type)
            else:
                logger.error(f"Unsupported file format: {file_extension}")
                return None
                
        except Exception as e:
            logger.error(f"Error parsing bank statement: {e}")
            return None
    
    def _parse_excel_statement(self, file_path: str, bank_type: str) -> pd.DataFrame:
        """Parse Excel bank statement"""
        try:
            # Read Excel file
            excel_file = pd.ExcelFile(file_path)
            
            # Try to find the right sheet
            sheet_name = None
            for sheet in excel_file.sheet_names:
                if any(keyword in sheet.lower() for keyword in ['transaction', 'statement', 'data']):
                    sheet_name = sheet
                    break
            
            if not sheet_name:
                sheet_name = excel_file.sheet_names[0]
            
            # Read the sheet
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Standardize column names based on bank type
            df = self._standardize_columns(df, bank_type)
            
            return df
            
        except Exception as e:
            logger.error(f"Error parsing Excel statement: {e}")
            raise
    
    def _parse_csv_statement(self, file_path: str, bank_type: str) -> pd.DataFrame:
        """Parse CSV bank statement"""
        try:
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'cp1252']
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    df = self._standardize_columns(df, bank_type)
                    return df
                except UnicodeDecodeError:
                    continue
            
            raise ValueError("Could not decode CSV file with any encoding")
            
        except Exception as e:
            logger.error(f"Error parsing CSV statement: {e}")
            raise
    
    def _parse_pdf_statement(self, file_path: str, bank_type: str) -> pd.DataFrame:
        """Parse PDF bank statement"""
        try:
            import PyPDF2
            import tabula
            
            # Try using tabula for table extraction
            try:
                tables = tabula.read_pdf(file_path, pages='all')
                if tables:
                    df = tables[0]  # Use first table
                    df = self._standardize_columns(df, bank_type)
                    return df
            except:
                pass
            
            # Fallback to text extraction
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text()
            
            # Parse text into structured data
            df = self._parse_text_to_dataframe(text, bank_type)
            return df
            
        except Exception as e:
            logger.error(f"Error parsing PDF statement: {e}")
            raise
    
    def _standardize_columns(self, df: pd.DataFrame, bank_type: str) -> pd.DataFrame:
        """Standardize column names for different banks"""
        # Common column mappings
        column_mappings = {
            'cbe': {
                'transaction_id': ['Transaction ID', 'STN', 'Reference', 'Ref No'],
                'amount': ['Amount', 'Transaction Amount', 'Debit', 'Credit'],
                'date': ['Date', 'Transaction Date', 'Value Date'],
                'sender': ['From', 'Sender', 'Account Name'],
                'receiver': ['To', 'Receiver', 'Beneficiary'],
                'description': ['Description', 'Narration', 'Remarks']
            },
            'telebirr': {
                'transaction_id': ['Transaction ID', 'STN', 'Reference'],
                'amount': ['Amount', 'Transaction Amount'],
                'date': ['Date', 'Transaction Date'],
                'sender': ['From', 'Sender'],
                'receiver': ['To', 'Receiver'],
                'description': ['Description', 'Purpose']
            },
            'dashen': {
                'transaction_id': ['Transaction ID', 'STN', 'Reference'],
                'amount': ['Amount', 'Transaction Amount'],
                'date': ['Date', 'Transaction Date'],
                'sender': ['From', 'Sender'],
                'receiver': ['To', 'Receiver'],
                'description': ['Description', 'Narration']
            }
        }
        
        mapping = column_mappings.get(bank_type, column_mappings['cbe'])
        
        # Rename columns
        for standard_name, possible_names in mapping.items():
            for col in df.columns:
                if any(name.lower() in col.lower() for name in possible_names):
                    df = df.rename(columns={col: standard_name})
                    break
        
        return df
    
    def _parse_text_to_dataframe(self, text: str, bank_type: str) -> pd.DataFrame:
        """Parse text content into structured DataFrame"""
        # This is a simplified parser - in practice, you'd use more sophisticated NLP
        lines = text.split('\n')
        data = []
        
        for line in lines:
            # Look for transaction patterns
            if any(keyword in line.lower() for keyword in ['transaction', 'payment', 'transfer']):
                # Extract transaction data using regex
                import re
                
                # Extract amount
                amount_match = re.search(r'([0-9,]+\.?[0-9]*)', line)
                amount = float(amount_match.group(1).replace(',', '')) if amount_match else 0
                
                # Extract date
                date_match = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', line)
                date = date_match.group(1) if date_match else None
                
                # Extract transaction ID
                id_match = re.search(r'([A-Z0-9]{8,})', line)
                transaction_id = id_match.group(1) if id_match else None
                
                if amount > 0 and transaction_id:
                    data.append({
                        'transaction_id': transaction_id,
                        'amount': amount,
                        'date': date,
                        'description': line.strip()
                    })
        
        return pd.DataFrame(data)
    
    def _perform_reconciliation(self, transactions_df: pd.DataFrame, 
                              statement_df: pd.DataFrame, 
                              bank_type: str) -> Dict:
        """Perform reconciliation between transactions and statement"""
        try:
            reconciliation_results = {
                'matched': [],
                'unmatched_transactions': [],
                'unmatched_statement': [],
                'discrepancies': []
            }
            
            # Clean and prepare data
            transactions_df = self._clean_transaction_data(transactions_df)
            statement_df = self._clean_statement_data(statement_df)
            
            # Match transactions with statement entries
            for _, transaction in transactions_df.iterrows():
                match_found = False
                
                for _, statement_row in statement_df.iterrows():
                    # Check for matches based on multiple criteria
                    if self._is_transaction_match(transaction, statement_row, bank_type):
                        reconciliation_results['matched'].append({
                            'transaction_id': transaction.get('id'),
                            'stn_number': transaction.get('stn_number'),
                            'amount': transaction.get('amount'),
                            'statement_amount': statement_row.get('amount'),
                            'transaction_date': transaction.get('transaction_date'),
                            'statement_date': statement_row.get('date'),
                            'confidence': self._calculate_match_confidence(transaction, statement_row)
                        })
                        match_found = True
                        break
                
                if not match_found:
                    reconciliation_results['unmatched_transactions'].append({
                        'transaction_id': transaction.get('id'),
                        'stn_number': transaction.get('stn_number'),
                        'amount': transaction.get('amount'),
                        'transaction_date': transaction.get('transaction_date'),
                        'waiter_id': transaction.get('waiter_id')
                    })
            
            # Find unmatched statement entries
            matched_statement_ids = [match['stn_number'] for match in reconciliation_results['matched']]
            
            for _, statement_row in statement_df.iterrows():
                if statement_row.get('transaction_id') not in matched_statement_ids:
                    reconciliation_results['unmatched_statement'].append({
                        'transaction_id': statement_row.get('transaction_id'),
                        'amount': statement_row.get('amount'),
                        'date': statement_row.get('date'),
                        'description': statement_row.get('description', '')
                    })
            
            # Identify discrepancies
            reconciliation_results['discrepancies'] = self._identify_discrepancies(
                reconciliation_results['matched']
            )
            
            return reconciliation_results
            
        except Exception as e:
            logger.error(f"Error in reconciliation: {e}")
            raise
    
    def _clean_transaction_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean transaction data for reconciliation"""
        # Remove duplicates
        df = df.drop_duplicates(subset=['stn_number'])
        
        # Convert amount to numeric
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
        
        # Convert date to datetime
        df['transaction_date'] = pd.to_datetime(df['transaction_date'], errors='coerce')
        
        # Remove rows with missing critical data
        df = df.dropna(subset=['stn_number', 'amount'])
        
        return df
    
    def _clean_statement_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean statement data for reconciliation"""
        # Remove duplicates
        df = df.drop_duplicates(subset=['transaction_id'])
        
        # Convert amount to numeric
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
        
        # Convert date to datetime
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        
        # Remove rows with missing critical data
        df = df.dropna(subset=['transaction_id', 'amount'])
        
        return df
    
    def _is_transaction_match(self, transaction: pd.Series, 
                            statement_row: pd.Series, 
                            bank_type: str) -> bool:
        """Check if transaction matches statement entry"""
        # Check STN number match
        if transaction.get('stn_number') == statement_row.get('transaction_id'):
            return True
        
        # Check amount match (with tolerance)
        amount_diff = abs(transaction.get('amount', 0) - statement_row.get('amount', 0))
        if amount_diff <= 1.0:  # 1 Birr tolerance
            # Check date proximity
            trans_date = transaction.get('transaction_date')
            stmt_date = statement_row.get('date')
            
            if trans_date and stmt_date:
                date_diff = abs((trans_date - stmt_date).days)
                if date_diff <= 3:  # 3 days tolerance
                    return True
        
        return False
    
    def _calculate_match_confidence(self, transaction: pd.Series, 
                                  statement_row: pd.Series) -> float:
        """Calculate confidence score for transaction match"""
        confidence = 0.0
        
        # STN number match (highest weight)
        if transaction.get('stn_number') == statement_row.get('transaction_id'):
            confidence += 0.8
        
        # Amount match
        amount_diff = abs(transaction.get('amount', 0) - statement_row.get('amount', 0))
        if amount_diff == 0:
            confidence += 0.2
        elif amount_diff <= 1.0:
            confidence += 0.1
        
        # Date match
        trans_date = transaction.get('transaction_date')
        stmt_date = statement_row.get('date')
        
        if trans_date and stmt_date:
            date_diff = abs((trans_date - stmt_date).days)
            if date_diff == 0:
                confidence += 0.1
            elif date_diff <= 1:
                confidence += 0.05
        
        return min(confidence, 1.0)
    
    def _identify_discrepancies(self, matched_transactions: List[Dict]) -> List[Dict]:
        """Identify discrepancies in matched transactions"""
        discrepancies = []
        
        for match in matched_transactions:
            # Check amount discrepancies
            amount_diff = abs(match['amount'] - match['statement_amount'])
            if amount_diff > 1.0:
                discrepancies.append({
                    'type': 'amount_mismatch',
                    'transaction_id': match['transaction_id'],
                    'stn_number': match['stn_number'],
                    'transaction_amount': match['amount'],
                    'statement_amount': match['statement_amount'],
                    'difference': amount_diff
                })
            
            # Check date discrepancies
            if match['transaction_date'] and match['statement_date']:
                date_diff = abs((match['transaction_date'] - match['statement_date']).days)
                if date_diff > 1:
                    discrepancies.append({
                        'type': 'date_mismatch',
                        'transaction_id': match['transaction_id'],
                        'stn_number': match['stn_number'],
                        'transaction_date': match['transaction_date'],
                        'statement_date': match['statement_date'],
                        'difference_days': date_diff
                    })
        
        return discrepancies
    
    def _generate_reconciliation_summary(self, reconciliation_results: Dict) -> Dict:
        """Generate summary statistics for reconciliation"""
        total_transactions = len(reconciliation_results['matched']) + len(reconciliation_results['unmatched_transactions'])
        total_statement_entries = len(reconciliation_results['matched']) + len(reconciliation_results['unmatched_statement'])
        
        matched_amount = sum(match['amount'] for match in reconciliation_results['matched'])
        unmatched_amount = sum(trans['amount'] for trans in reconciliation_results['unmatched_transactions'])
        
        return {
            'total_transactions': total_transactions,
            'total_statement_entries': total_statement_entries,
            'matched_transactions': len(reconciliation_results['matched']),
            'unmatched_transactions': len(reconciliation_results['unmatched_transactions']),
            'unmatched_statement_entries': len(reconciliation_results['unmatched_statement']),
            'matched_amount': matched_amount,
            'unmatched_amount': unmatched_amount,
            'discrepancies': len(reconciliation_results['discrepancies']),
            'match_rate': len(reconciliation_results['matched']) / total_transactions if total_transactions > 0 else 0
        }
    
    def generate_audit_report(self, reconciliation_results: Dict, 
                            summary: Dict, 
                            report_date: datetime,
                            admin_name: str) -> str:
        """Generate audit report in specified format"""
        try:
            timestamp = report_date.strftime('%Y%m%d_%H%M%S')
            
            if self.report_format == 'pdf':
                return self._generate_pdf_report(reconciliation_results, summary, report_date, admin_name, timestamp)
            elif self.report_format == 'excel':
                return self._generate_excel_report(reconciliation_results, summary, report_date, admin_name, timestamp)
            elif self.report_format == 'csv':
                return self._generate_csv_report(reconciliation_results, summary, report_date, admin_name, timestamp)
            else:
                raise ValueError(f"Unsupported report format: {self.report_format}")
                
        except Exception as e:
            logger.error(f"Error generating audit report: {e}")
            raise
    
    def _generate_pdf_report(self, reconciliation_results: Dict, summary: Dict,
                           report_date: datetime, admin_name: str, timestamp: str) -> str:
        """Generate PDF audit report"""
        filename = f"audit_report_{timestamp}.pdf"
        filepath = self.reports_dir / filename
        
        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        story = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        story.append(Paragraph("VeriPay Audit Report", title_style))
        story.append(Spacer(1, 20))
        
        # Summary information
        story.append(Paragraph(f"<b>Report Date:</b> {report_date.strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Paragraph(f"<b>Generated By:</b> {admin_name}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Summary table
        summary_data = [
            ['Metric', 'Value'],
            ['Total Transactions', summary['total_transactions']],
            ['Matched Transactions', summary['matched_transactions']],
            ['Unmatched Transactions', summary['unmatched_transactions']],
            ['Match Rate', f"{summary['match_rate']:.2%}"],
            ['Matched Amount', f"ETB {summary['matched_amount']:,.2f}"],
            ['Unmatched Amount', f"ETB {summary['unmatched_amount']:,.2f}"],
            ['Discrepancies', summary['discrepancies']]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Unmatched transactions table
        if reconciliation_results['unmatched_transactions']:
            story.append(Paragraph("<b>Unmatched Transactions</b>", styles['Heading2']))
            story.append(Spacer(1, 10))
            
            unmatched_data = [['STN Number', 'Amount', 'Date', 'Waiter ID']]
            for trans in reconciliation_results['unmatched_transactions']:
                unmatched_data.append([
                    trans['stn_number'],
                    f"ETB {trans['amount']:,.2f}",
                    trans['transaction_date'].strftime('%Y-%m-%d') if trans['transaction_date'] else 'N/A',
                    str(trans['waiter_id'])
                ])
            
            unmatched_table = Table(unmatched_data)
            unmatched_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(unmatched_table)
        
        doc.build(story)
        return str(filepath)
    
    def _generate_excel_report(self, reconciliation_results: Dict, summary: Dict,
                             report_date: datetime, admin_name: str, timestamp: str) -> str:
        """Generate Excel audit report"""
        filename = f"audit_report_{timestamp}.xlsx"
        filepath = self.reports_dir / filename
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add summary information
        ws_summary['A1'] = "VeriPay Audit Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        ws_summary['A3'] = f"Report Date: {report_date.strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary['A4'] = f"Generated By: {admin_name}"
        
        # Add summary table
        summary_data = [
            ['Metric', 'Value'],
            ['Total Transactions', summary['total_transactions']],
            ['Matched Transactions', summary['matched_transactions']],
            ['Unmatched Transactions', summary['unmatched_transactions']],
            ['Match Rate', f"{summary['match_rate']:.2%}"],
            ['Matched Amount', summary['matched_amount']],
            ['Unmatched Amount', summary['unmatched_amount']],
            ['Discrepancies', summary['discrepancies']]
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Style summary table
        for cell in ws_summary['A6:H6']:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Matched transactions sheet
        if reconciliation_results['matched']:
            ws_matched = wb.create_sheet("Matched Transactions")
            matched_data = [['STN Number', 'Amount', 'Statement Amount', 'Transaction Date', 'Statement Date', 'Confidence']]
            
            for match in reconciliation_results['matched']:
                matched_data.append([
                    match['stn_number'],
                    match['amount'],
                    match['statement_amount'],
                    match['transaction_date'].strftime('%Y-%m-%d') if match['transaction_date'] else 'N/A',
                    match['statement_date'].strftime('%Y-%m-%d') if match['statement_date'] else 'N/A',
                    f"{match['confidence']:.2%}"
                ])
            
            for row in matched_data:
                ws_matched.append(row)
            
            # Style header
            for cell in ws_matched['A1:F1']:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Unmatched transactions sheet
        if reconciliation_results['unmatched_transactions']:
            ws_unmatched = wb.create_sheet("Unmatched Transactions")
            unmatched_data = [['STN Number', 'Amount', 'Transaction Date', 'Waiter ID']]
            
            for trans in reconciliation_results['unmatched_transactions']:
                unmatched_data.append([
                    trans['stn_number'],
                    trans['amount'],
                    trans['transaction_date'].strftime('%Y-%m-%d') if trans['transaction_date'] else 'N/A',
                    trans['waiter_id']
                ])
            
            for row in unmatched_data:
                ws_unmatched.append(row)
            
            # Style header
            for cell in ws_unmatched['A1:D1']:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        wb.save(str(filepath))
        return str(filepath)
    
    def _generate_csv_report(self, reconciliation_results: Dict, summary: Dict,
                           report_date: datetime, admin_name: str, timestamp: str) -> str:
        """Generate CSV audit report"""
        filename = f"audit_report_{timestamp}.csv"
        filepath = self.reports_dir / filename
        
        # Create summary DataFrame
        summary_df = pd.DataFrame([
            ['Report Date', report_date.strftime('%Y-%m-%d %H:%M:%S')],
            ['Generated By', admin_name],
            ['Total Transactions', summary['total_transactions']],
            ['Matched Transactions', summary['matched_transactions']],
            ['Unmatched Transactions', summary['unmatched_transactions']],
            ['Match Rate', f"{summary['match_rate']:.2%}"],
            ['Matched Amount', summary['matched_amount']],
            ['Unmatched Amount', summary['unmatched_amount']],
            ['Discrepancies', summary['discrepancies']]
        ], columns=['Metric', 'Value'])
        
        # Create matched transactions DataFrame
        if reconciliation_results['matched']:
            matched_df = pd.DataFrame(reconciliation_results['matched'])
            matched_df['transaction_date'] = matched_df['transaction_date'].dt.strftime('%Y-%m-%d')
            matched_df['statement_date'] = matched_df['statement_date'].dt.strftime('%Y-%m-%d')
            matched_df['confidence'] = matched_df['confidence'].apply(lambda x: f"{x:.2%}")
        
        # Create unmatched transactions DataFrame
        if reconciliation_results['unmatched_transactions']:
            unmatched_df = pd.DataFrame(reconciliation_results['unmatched_transactions'])
            unmatched_df['transaction_date'] = unmatched_df['transaction_date'].dt.strftime('%Y-%m-%d')
        
        # Write to CSV
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            # Write summary
            f.write("=== SUMMARY ===\n")
            summary_df.to_csv(f, index=False)
            f.write("\n")
            
            # Write matched transactions
            if reconciliation_results['matched']:
                f.write("=== MATCHED TRANSACTIONS ===\n")
                matched_df.to_csv(f, index=False)
                f.write("\n")
            
            # Write unmatched transactions
            if reconciliation_results['unmatched_transactions']:
                f.write("=== UNMATCHED TRANSACTIONS ===\n")
                unmatched_df.to_csv(f, index=False)
        
        return str(filepath) 